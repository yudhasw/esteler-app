"""
Routes Admin - panel pengelolaan penjual
Semua route dilindungi @login_required
"""
from flask import (
    Blueprint, request, redirect, url_for, render_template, flash, jsonify, send_file
)
from flask_login import login_required, current_user

from services.menu_service import MenuService
from services.order_service import OrderService
from services.analytics_service import AnalyticsService
from models import Order
import io

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")


# ====================== DASHBOARD ======================
@admin_bp.route("/")
@admin_bp.route("/dashboard")
@login_required
def dashboard():
    summary = AnalyticsService.get_dashboard_summary()
    top_menus = AnalyticsService.get_top_menus(limit=3, days=30)
    weekly_revenue = AnalyticsService.get_weekly_revenue()
    recent_orders = OrderService.get_recent(limit=5)

    return render_template(
        "admin/dashboard.html",
        user=current_user,
        summary=summary,
        top_menus=top_menus,
        weekly_revenue=weekly_revenue,
        recent_orders=recent_orders,
    )


# ====================== MENU MANAGEMENT ======================
@admin_bp.route("/menu")
@login_required
def menu_management():
    menus = MenuService.get_all(active_only=False)
    summary = {
        "total": len(menus),
        "active": sum(1 for m in menus if m.is_active),
        "completed_all": OrderService.count_completed_all(),
    }
    return render_template("admin/menu_list.html", menus=menus, summary=summary)


@admin_bp.route("/menu/create", methods=["GET", "POST"])
@login_required
def menu_create():
    if request.method == "POST":
        try:
            MenuService.create({
                "name": request.form.get("name", "").strip(),
                "description": request.form.get("description", "").strip(),
                "price": int(request.form.get("price", 0)),
                "category": request.form.get("category", "Original"),
                "image_url": request.form.get("image_url", "").strip(),
                "is_active": request.form.get("is_active") == "on",
                "is_best_seller": request.form.get("is_best_seller") == "on",
                "is_favorite": request.form.get("is_favorite") == "on",
            })
            flash("Menu berhasil ditambahkan.", "success")
            return redirect(url_for("admin.menu_management"))
        except (ValueError, KeyError) as e:
            flash(f"Data tidak valid: {e}", "danger")
    return render_template("admin/menu_form.html", menu=None)


@admin_bp.route("/menu/<int:menu_id>/edit", methods=["GET", "POST"])
@login_required
def menu_edit(menu_id):
    menu = MenuService.get_by_id(menu_id)
    if not menu:
        flash("Menu tidak ditemukan.", "danger")
        return redirect(url_for("admin.menu_management"))

    if request.method == "POST":
        MenuService.update(menu_id, {
            "name": request.form.get("name", "").strip(),
            "description": request.form.get("description", "").strip(),
            "price": int(request.form.get("price", menu.price)),
            "category": request.form.get("category", menu.category),
            "image_url": request.form.get("image_url", menu.image_url),
            "is_active": request.form.get("is_active") == "on",
            "is_best_seller": request.form.get("is_best_seller") == "on",
            "is_favorite": request.form.get("is_favorite") == "on",
        })
        flash("Menu berhasil diperbarui.", "success")
        return redirect(url_for("admin.menu_management"))

    return render_template("admin/menu_form.html", menu=menu)


@admin_bp.route("/menu/<int:menu_id>/toggle", methods=["POST"])
@login_required
def toggle_menu(menu_id):
    if MenuService.toggle_active(menu_id):
        flash("Status menu berhasil diubah.", "success")
    else:
        flash("Menu tidak ditemukan.", "danger")
    return redirect(url_for("admin.menu_management"))


# ====================== ORDER MANAGEMENT ======================
@admin_bp.route("/orders")
@login_required
def order_list():
    active = OrderService.get_active()
    recent = OrderService.get_recent(limit=20)
    return render_template(
        "admin/orders.html", active_orders=active, recent_orders=recent
    )


@admin_bp.route("/orders/<int:order_id>/process", methods=["POST"])
@login_required
def process_order(order_id):
    try:
        OrderService.update_status(order_id, "memasak")
        flash("Pesanan diproses.", "success")
    except ValueError as e:
        flash(str(e), "danger")
    return redirect(request.referrer or url_for("admin.order_list"))


@admin_bp.route("/orders/<int:order_id>/ready", methods=["POST"])
@login_required
def ready_order(order_id):
    OrderService.update_status(order_id, "siap")
    flash("Pesanan siap diambil.", "success")
    return redirect(request.referrer or url_for("admin.order_list"))


@admin_bp.route("/orders/<int:order_id>/complete", methods=["POST"])
@login_required
def complete_order(order_id):
    OrderService.update_status(order_id, "selesai")
    flash("Pesanan selesai.", "success")
    return redirect(request.referrer or url_for("admin.order_list"))


@admin_bp.route("/orders/<int:order_id>/cancel", methods=["POST"])
@login_required
def cancel_order(order_id):
    OrderService.cancel(order_id)
    flash("Pesanan dibatalkan.", "warning")
    return redirect(request.referrer or url_for("admin.order_list"))


# ====================== WALK-IN ORDER (sesuai SDD revisi) ======================
@admin_bp.route("/orders/walkin", methods=["GET", "POST"])
@login_required
def walkin_order():
    if request.method == "POST":
        # Format: menu_id[]=1&quantity[]=2&menu_id[]=3&quantity[]=1
        menu_ids = request.form.getlist("menu_id[]")
        quantities = request.form.getlist("quantity[]")

        items = [
            {"menu_id": int(mid), "quantity": int(q)}
            for mid, q in zip(menu_ids, quantities)
            if mid and int(q) > 0
        ]

        if not items:
            flash("Tidak ada item yang dipilih.", "danger")
            return redirect(url_for("admin.walkin_order"))

        order = OrderService.create_walkin(items, {
            "name": request.form.get("customer_name", "Walk-in"),
            "contact": request.form.get("customer_contact"),
            "note": request.form.get("note", ""),
            "payment_method": request.form.get("payment_method", "TUNAI"),
        })

        if order:
            flash(f"Order walk-in {order.order_code} berhasil dibuat.", "success")
            return redirect(url_for("admin.order_list"))
        flash("Gagal membuat order.", "danger")

    menus = MenuService.get_all(active_only=True)
    return render_template("admin/walkin.html", menus=menus)

# ====================== REPORT DOWNLOAD ======================
@admin_bp.route("/report/download")
@login_required
def download_report():
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # 1. Ambil data order terbaru
    orders = Order.query.order_by(Order.created_at.desc()).all()
    
    # 2. Buat workbook excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Penjualan"
    
    # Header kolom
    headers = [
        "Tanggal", "Kode Pesanan", "Nama Pelanggan", "Kontak", 
        "Tipe Order", "Sumber", "Status", "Metode Bayar", "Menu Dipesan", "Total (Rp)"
    ]
    ws.append(headers)
    
    # Styling header (Warna Hijau Brand)
    header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
    
    # Isi data baris per baris
    for o in orders:
        menu_details = ", ".join([f"{item.menu.name if item.menu else 'Menu'} ({item.quantity}x)" for item in o.items])
        
        # Jika pesanan belum selesai, pendapatannya 0
        revenue = o.total if o.status and o.status.lower() == "selesai" else 0
        
        ws.append([
            o.created_at.strftime("%Y-%m-%d %H:%M") if o.created_at else "-",
            o.order_code,
            o.customer_name or "-",
            o.customer_contact or "-",
            o.order_type.upper() if o.order_type else "-",
            o.order_source,
            o.status.upper(),
            o.payment_method,
            menu_details,
            revenue
        ])
        
    # Tambahkan baris Total Pendapatan di bawah
    # Kita hitung langsung di Python agar nilainya pasti muncul tanpa perlu bergantung pada kalkulasi otomatis Excel
    total_pendapatan = sum(o.total for o in orders if o.status and o.status.lower() == "selesai")
    last_data_row = len(orders) + 1
    
    ws.append(["", "", "", "", "", "", "", "", "Total Pendapatan:", total_pendapatan])
    
    total_row_idx = last_data_row + 1
    ws.cell(row=total_row_idx, column=9).font = Font(bold=True)
    ws.cell(row=total_row_idx, column=10).font = Font(bold=True)
        
    # Auto-adjust lebar kolom biar rapi
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # A, B, C, dst.
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # Tambah margin sedikit
        ws.column_dimensions[column].width = max_length + 2
        
    # Simpan file ke in-memory buffer (tidak perlu save ke harddisk)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Kirim file ke browser
    return send_file(
        output,
        download_name="Laporan_Penjualan_Dapur_Hijrah.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
